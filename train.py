import matplotlib.pyplot as plt
from dataloader import get_loader
from encoder import *
import torch.optim as optim
import openpyxl

#used for training the model

workbook = openpyxl.Workbook()
sheet = workbook.active

def main():
    Net = Generator()
    ResNet = make_model()

    device = torch.device("cuda:0")  # 指定要使用的 CUDA 设备
    Net.to(device)

    ResNet.to(device)

    loss_function = nn.MSELoss()

    num_epoch = 150

    train_loss_list = list()
    val_loss_list = list()
    optimizer = optim.Adam(Net.parameters(), lr=1e-4)

    data_loader = get_loader(root = r'train', batch_size=1024, num_works=0)
    val_dataloader = get_loader(root = r'test', batch_size=1024, num_works=0)
    min_val = 100
    for epoch in range(num_epoch):
        trn = []
        Net.train()
        for iterno, s_t in enumerate(data_loader):
            s_t = s_t.to(device)
            x_t, x_pred_t  = Net(s_t)
            loss = torch.sqrt(loss_function(x_pred_t.to(device), x_t.to(device)))
            trn.append(loss.item())
            optimizer.zero_grad()
            loss.backward()
            torch.nn.utils.clip_grad_value_(Net.parameters(), 10)
            optimizer.step()

        train_loss = (sum(trn) / len(trn))
        train_loss_list.append(train_loss)

        Net.eval()
        val = []
        with torch.no_grad():
            for s_t in val_dataloader:
                s_t = s_t.detach()
                x_pred_t, x_t = Net(s_t.cuda())
                loss = torch.sqrt(loss_function(x_pred_t.cuda(), x_t.cuda()))
                val.append(loss)

        val_loss = (sum(val) / len(val)).item()
        val_loss_list.append(val_loss)
        print('epoch : ', epoch,  ' | TL : ', train_loss, ' | VL : ', val_loss)


        if val_loss < min_val:
            print('saving model')
            min_val = val_loss
            torch.save(Net.state_dict(), 'model1/model%s.pt'%epoch) #保存的模型对应的路径

        if (epoch+1) % 100 == 0:
            print('saving model')
            torch.save(Net.state_dict(), 'model1/model%s.pt' % epoch) #保存的模型对应的路径

    for j in range(len(train_loss_list)):
        sheet.cell(row=j + 1, column=1, value=train_loss_list[j])
    workbook.save(r"model1/train_loss.xlsx")

    for j in range(len(val_loss_list)):
        sheet.cell(row=j + 1, column=1, value=val_loss_list[j])
    workbook.save(r"model1/val_loss.xlsx")

    plt.plot(train_loss_list)
    plt.plot(val_loss_list)
    plt.show()

if __name__ == "__main__":
    main()
